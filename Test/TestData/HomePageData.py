import openpyxl

class HomePageData:
    
    test_HomePage_data = [{"name":"Alejandra","email":"alejandra.evander92@gmail.com","gender":"Male","password":"12345678i"},{"name":"Baka","email":"bakamoto@gmail.com","gender":"Female","password":"12345678o"}]
    
    @staticmethod
    def getTestData(test_Case_name):
        Dict={}
        book = openpyxl.load_workbook(r"C:\Users\User\Documents\Details.xlsx")
        sheet = book.active
        for i in range(1,sheet.max_row+1):
            if sheet.cell(row=i, column=1).value==test_Case_name:
                
                for j in range(2, sheet.max_column+1):
                    Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value
        print(Dict)